#login page GUI
from tkinter import *
import Signup
import Book
import openpyxl as xl
dic={}
gf=xl.load_workbook('DBMS/base2.xlsx')
sh=gf.active
for i in range(2,sh.max_row):
    reg=sh.cell(i,2).value
    passw=sh.cell(i,5).value
    dic[reg]=passw




print(dic)



if __name__=="__main__":
    def clear(): # clearing the content after storing
        reg_no_field.delete(0, END) 
        password_field.delete(0, END)
        reg_no_field.focus_set()
    def valid():
        try:
            
            a=reg_no_field.get()
            b=password_field.get()
            if  a not in dic.keys():
                print("Username doesnt exist")
                clear()
            elif b!=dic[a]:
                print("Incorrect paasword")
                clear()
            else:
                Book.main(reg_no_field.get())
                

        except ValueError:
            print("Enter correct Values")
            clear()
                
            
        


    
    def focus1(event):
        password_field.focus_set() 
   
  

    root = Tk()
    root.configure(background='Grey') 
    root.title("Login page") 
    root.geometry("500x300") 
   
  
    #Login label 
    heading = Label(root, text="LOGIN", bg="Grey", fg="black") 
    heading.grid(row=0, column=1)
    #Name label 
    reg_no = Label(root, text="Registration Number", bg="Grey", fg="White") 
    reg_no.grid(row=1, column=0) 
    #password label 
    password = Label(root, text="Password", bg="Grey", fg="White")
    password.grid(row=2, column=0) 
    
  
    #entry box 
    reg_no_field = Entry(root) 
    password_field = Entry(root) 
    
    reg_no_field.bind("<Return>", focus1)  
  
     
    reg_no_field.grid(row=1, column=1, ipadx="50") 
    password_field.grid(row=2, column=1, ipadx="50") 
     
 
    #Login button 
    Login = Button(root, text="Login", fg="white", 
                            bg="orange", command=valid) 
    Login.grid(row=8, column=1)

    #Signup Button
    signup = Button(root, text="SignUp", fg="white", 
                            bg="orange", command=Signup.main2) 
    signup.grid(row=9, column=1) 
  
    # start the GUI 
    root.mainloop()


