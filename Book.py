
import tkinter as tk
from tkinter import ttk

from tkcalendar import Calendar, DateEntry
#Database part
from openpyxl import *
sb=load_workbook('Database1.xlsx')
sheet=sb.active
#database part end
fu2=""



def main(R):
    d=R+" Book A Ride"
    def store():
        if (t.get()==""): 
              print("empty Time") 
        else:
    
            cr = sheet.max_row
            sheet.cell(row=cr + 1, column=1).value = w.get() 
            sheet.cell(row=cr + 1, column=2).value = y.get() 
            sheet.cell(row=cr + 1, column=3).value = R
            sheet.cell(row=cr + 1, column=4).value = fu2
            sheet.cell(row=cr + 1, column=5).value = t.get()
            sb.save('Database1.xlsx')
            main = tk.Tk() 
            ourMessage ='Your ride will arrive in a minute'
            messageVar = tk.Message(main, text = ourMessage) 
            messageVar.pack( ) 
            main.mainloop( ) 
    def example1():
        '''def doo():
            cr = sheet.max_row
            sheet.cell(row=cr + 1, column=1).value = w.get() 
            sheet.cell(row=cr + 1, column=2).value = y.get() 
            sheet.cell(row=cr + 1, column=3).value = R
            sheet.cell(row=cr + 1, column=4).value = cal.selection_get()
            sheet.cell(row=cr + 1, column=5).value = t.get()
            fb.save('Database1.xlsx')
            main = tk.Tk() 
            ourMessage ='Your ride will arrive in minute'
            messageVar = Message(main, text = ourMessage) 
            messageVar.pack( ) 
            main.mainloop( ) '''
        def fu():
            global fu2
            fu2=cal.selection_get()
        top = tk.Toplevel(root)

        cal = Calendar(top,
                   font="Arial 14", selectmode='day',
                   cursor="hand1", year=2019, month=11, day=5)
        cal.pack(fill="both", expand=True)
        ttk.Button(top, text="ok", command=fu).pack()

    
    root =tk.Tk()
    s = ttk.Style(root)
    s.theme_use('clam')
    root.configure(background='#ffde85') 
    root.title("Book A Ride") 
    root.geometry("500x300") 
  
    #Login label 
    heading = tk.Label(root, text=d, bg="#ffde85", fg="black") 
    heading.grid(row=0, column=1)
    #Name label 
    From = tk.Label(root, text="From Block No:", bg="#ffde85", fg="White")
    From.grid(row=1, column=0) 
    #password label 
    To = tk.Label(root, text="To Block No:", bg="#ffde85", fg="White")
    To.grid(row=2, column=0)
    Time = tk.Label(root, text="Time(HH:MM)", bg="#ffde85", fg="White").grid(row=3,column=0)
    w=tk.Spinbox(root, from_ = 1, to = 58)  
    y=tk.Spinbox(root, from_ = 1, to = 58)
    t=tk.Entry(root)
    t.grid(row=3, column=1)
    w.grid(row=1, column=1, ipadx="50") 
    y.grid(row=2, column=1, ipadx="50")
   # tim=TimeEntry(top, width=12, background='darkblue',
                   #5 foreground='white', borderwidth=2)
    ttk.Button(root, text='Date', command=example1).grid(row=4,column=1)
    
 
    #Login button 
    Book = tk.Button(root, text="Book", fg="white", 
                            bg="orange", command=store) 
    Book.grid(row=5, column=1)
    root.mainloop()

