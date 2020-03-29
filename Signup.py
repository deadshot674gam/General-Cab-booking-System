from tkinter import *

#Database part
from openpyxl import *
fb=load_workbook('DBMS/base2.xlsx')
sheet=fb.active
'''
def excel(): 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
  
    
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Registration Number"
    sheet.cell(row=1, column=3).value = "Mobile Number"
    sheet.cell(row=1, column=4).value = "Email-Id"
    sheet.cell(row=1, column=5).value = "Password"
#database part end

'''

#main function for signup form
def main2():


    
    def focus1(event):  
        Reg_field.focus_set() 
    def focus2(event): 
        Mob_field.focus_set() 
    def focus3(event): 
        email_id_field.focus_set()
    def focus4(event): 
        passwor_field.focus_set()
        


    def clear(): # clearing the content after storing
        name_field.delete(0, END) 
        Reg_field.delete(0, END) 
        Mob_field.delete(0, END) 
        email_id_field.delete(0, END) 
        passwor_field.delete(0, END)
    def insert():  #updating the data base
        if (name_field.get() == "" and Reg_field.get() == "" and Mob_field.get() == "" and email_id_field.get() == "" and passwor_field==""): 
              print("empty input") 
        else: 
            cr = sheet.max_row
            sheet.cell(row=cr + 1, column=1).value = name_field.get() 
            sheet.cell(row=cr + 1, column=2).value = Reg_field.get() 
            sheet.cell(row=cr + 1, column=3).value = Mob_field.get() 
            sheet.cell(row=cr + 1, column=4).value = email_id_field.get() 
            sheet.cell(row=cr + 1, column=5).value = passwor_field.get()
            fb.save('DBMS/base2.xlsx')
            name_field.focus_set()
            clear() 


    
    # create a GUI window 
    root = Tk()
    root.configure(background='light green')
    root.title("SignUP Details") 
    root.geometry("500x300")
    
    #Label
    heading = Label(root, text="SignUP", bg="light green")
    name = Label(root, text="Name", bg="light green")
    Reg_no = Label(root, text="Registration Number", bg="light green") 
    Mob_no = Label(root, text="Mobile No:", bg="light green") 
    email_id = Label(root, text="Email id", bg="light green")
    passwor=Label(root, text="New Password", bg="light green")

    #grids for Label
    heading.grid(row=0, column=1) 
    name.grid(row=1, column=0) 
    Reg_no.grid(row=2, column=0) 
    Mob_no.grid(row=3, column=0) 
    email_id.grid(row=4, column=0)
    passwor.grid(row=5,column=0)
  
    #Entry
    name_field = Entry(root) 
    Reg_field = Entry(root) 
    Mob_field = Entry(root) 
    email_id_field = Entry(root) 
    passwor_field=Entry(root)
    
    #bind
    name_field.bind("<Return>", focus1)
    Reg_field.bind("<Return>", focus2)
    Mob_field.bind("<Return>", focus3)
    email_id_field.bind("<Return>", focus4)
    
  
    #grids for Entry
    name_field.grid(row=1, column=1, ipadx="100") 
    Reg_field.grid(row=2, column=1, ipadx="100") 
    Mob_field.grid(row=3, column=1, ipadx="100")  
    email_id_field.grid(row=4, column=1, ipadx="100") 
    passwor_field.grid(row=5,column=1,ipadx="100")

    
    submit = Button(root, text="SignUp", fg="Black", bg="Red", command=insert) 
    submit.grid(row=8, column=1)
    root.mainloop()
    

